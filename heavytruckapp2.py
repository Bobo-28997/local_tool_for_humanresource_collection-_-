# =====================================
# Tkinter Local App: 模拟人事用合同记录表自动审核
# (V5: 高效模式重构)
#
# - [新] 重构: 主文件 '月重卡' 现在只会被读取一次, 大幅提升性能
# - [新] "起租" sheet 已添加
# - [新] 切换到 Calamine 引擎, 忽略 Excel 筛选器错误
# - 保留 AuditApp GUI 模板、线程管理、TEMP 文件清理
# - 保留所有 "App 1" 的核心业务逻辑
# =====================================
import pandas as pd
import time
import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import threading
import traceback


# =====================================
# 🧹 [V3 功能] 启动时自动清理缓存
# =====================================
def clear_local_cache_on_startup():
    """
    (安全版本) 清理上次运行崩溃时残留的 TEMP_ 文件。
    """
    print("--- 启动维护：正在扫描旧的临时文件... ---")
    cleaned_count = 0
    try:
        current_dir = os.getcwd()
        for f_name in os.listdir(current_dir):
            if f_name.startswith("TEMP__") and f_name.endswith(".xlsx"):
                try:
                    os.remove(os.path.join(current_dir, f_name))
                    print(f"  > 已清理: {f_name}")
                    cleaned_count += 1
                except Exception as e:
                    print(f"  > ⚠️ 清理失败 (可能被占用): {f_name}. 错误: {e}")
        print(f"--- 维护完成：共清理 {cleaned_count} 个文件。 ---")
    except Exception as e:
        print(f"--- ⚠️ 启动维护失败: {e} ---")


# =====================================
# 🖥️ Tkinter 应用主类 (标准模板)
# =====================================
class AuditApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📊 人事薪资表自动审核系统-1 (重卡-高效版)")  # <--- (标题更新)
        self.root.geometry("800x700")

        self.style = ttk.Style()
        self.style.theme_use('clam')

        # --- 实例变量 ---
        self.uploaded_files = {}
        self.output_dir = ""
        self.required_files = ["月重卡", "放款明细", "字段", "二次明细"]
        self.all_files_to_save = []

        # --- GUI 布局 (标准模板) ---
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 1. 文件选择区
        input_frame = ttk.LabelFrame(main_frame, text="1. 输入文件", padding="10")
        input_frame.pack(fill=tk.X, expand=False)
        self.select_files_button = ttk.Button(input_frame, text=f"选择 {len(self.required_files)} 个 Excel 文件",
                                              command=self.select_files)
        self.select_files_button.pack(side=tk.LEFT, padx=(0, 10))
        self.file_status_label = ttk.Label(input_frame, text="尚未选择文件")
        self.file_status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # 2. 输出文件夹选择区
        output_frame = ttk.LabelFrame(main_frame, text="2. 输出文件夹", padding="10")
        output_frame.pack(fill=tk.X, expand=False, pady=5)
        self.select_output_dir_button = ttk.Button(output_frame, text="选择报告保存位置",
                                                   command=self.select_output_dir)
        self.select_output_dir_button.pack(side=tk.LEFT, padx=(0, 10))
        self.output_dir_label = ttk.Label(output_frame, text="尚未选择文件夹")
        self.output_dir_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # 3. 执行区
        run_frame = ttk.Frame(main_frame, padding="10")
        run_frame.pack(fill=tk.X, expand=False)
        self.run_button = ttk.Button(run_frame, text="🚀 开始审核", command=self.run_audit_thread, state=tk.DISABLED)
        self.run_button.pack(fill=tk.X, expand=True)

        # 4. 状态和进度区
        status_frame = ttk.LabelFrame(main_frame, text="3. 审核状态", padding="10")
        status_frame.pack(fill=tk.X, expand=False, pady=5)
        self.status_label = ttk.Label(status_frame, text="等待开始...")
        self.status_label.pack(fill=tk.X, expand=True)
        self.progress_bar = ttk.Progressbar(status_frame, orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress_bar.pack(fill=tk.X, expand=True, pady=5)

        # 5. 日志区
        log_frame = ttk.LabelFrame(main_frame, text="4. 运行日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.log_widget = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=20, state=tk.DISABLED)
        self.log_widget.pack(fill=tk.BOTH, expand=True)

        self.log_widget.tag_config('INFO', foreground='black')
        self.log_widget.tag_config('SUCCESS', foreground='green')
        self.log_widget.tag_config('WARNING', foreground='orange')
        self.log_widget.tag_config('ERROR', foreground='red', font=('Helvetica', '9', 'bold'))

    # =====================================
    # 🖥️ GUI 交互方法 (标准模板)
    # =====================================

    def _log(self, message, level='INFO'):
        """线程安全的日志更新"""
        self.root.after(0, self.update_log_widget, message, level)

    def update_log_widget(self, message, level):
        self.log_widget.config(state=tk.NORMAL)
        self.log_widget.insert(tk.END, f"{message}\n", level.upper())
        self.log_widget.see(tk.END)
        self.log_widget.config(state=tk.DISABLED)

    def _update_status(self, text):
        """线程安全的状态标签更新"""
        self.root.after(0, self.status_label.config, {'text': text})

    def _update_progress(self, value):
        """线程安全的进度条更新 (value: 0.0 to 1.0)"""
        self.root.after(0, self.progress_bar.config, {'value': value * 100})

    def _set_gui_state(self, is_running):
        """线程安全地切换按钮状态"""
        self.root.after(0, self.toggle_buttons, is_running)

    def toggle_buttons(self, is_running):
        state = tk.DISABLED if is_running else tk.NORMAL
        self.select_files_button.config(state=state)
        self.select_output_dir_button.config(state=state)
        if not is_running and self.check_ready(silent=True):
            self.run_button.config(state=tk.NORMAL)
        else:
            self.run_button.config(state=tk.DISABLED)

    def check_ready(self, silent=False):
        """检查所有条件是否满足，以启用“开始审核”按钮"""
        ready = len(self.uploaded_files) == len(self.required_files) and self.output_dir
        if ready:
            self.run_button.config(state=tk.NORMAL)
            if not silent:
                self._log("✅ 所有文件和输出目录已就绪，可以开始审核。", "SUCCESS")
        else:
            self.run_button.config(state=tk.DISABLED)
        return ready

    def select_files(self):
        self.uploaded_files = {}

        file_paths = filedialog.askopenfilenames(
            title=f"请选择所有 {len(self.required_files)} 个文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if not file_paths:
            self.file_status_label.config(text="未选择文件")
            self.check_ready()
            return

        found_count = 0
        missing_files = list(self.required_files)

        for keyword in self.required_files:
            found_for_keyword = False
            for path in file_paths:
                filename = os.path.basename(path)
                if keyword in filename:
                    self.uploaded_files[keyword] = path
                    found_count += 1
                    if keyword in missing_files:
                        missing_files.remove(keyword)
                    found_for_keyword = True
                    break

        self.file_status_label.config(
            text=f"已选择 {found_count} / {len(self.required_files)} 个文件。"
        )

        if missing_files:
            self._log(f"⚠️ 仍缺少文件: {', '.join(missing_files)}", "WARNING")

        self.check_ready()

    def select_output_dir(self):
        dir_path = filedialog.askdirectory(title="选择报告保存的文件夹")
        if dir_path:
            self.output_dir = dir_path
            self.output_dir_label.config(text=f"将保存到: {dir_path}")
        else:
            self.output_dir = ""
            self.output_dir_label.config(text="尚未选择文件夹")
        self.check_ready()

    def run_audit_thread(self):
        """“开始审核”按钮的入口，启动一个新线程"""
        self._set_gui_state(is_running=True)
        self.log_widget.config(state=tk.NORMAL)
        self.log_widget.delete('1.0', tk.END)
        self.log_widget.config(state=tk.DISABLED)

        self.all_files_to_save = []

        self._log("🚀 审核任务开始...", "INFO")
        threading.Thread(target=self.run_audit_logic, daemon=True).start()

    # =====================================
    # 核心业务逻辑 (移植为类方法)
    # =====================================

    def normalize_contract_key(self, series: pd.Series) -> pd.Series:
        s = series.astype(str)
        s = s.str.replace(r"\.0$", "", regex=True)
        s = s.str.strip()
        s = s.str.upper()
        s = s.str.replace('－', '-', regex=False)
        s = s.str.replace(r'\s+', '', regex=True)
        return s

    def normalize_colname(self, c):
        return str(c).strip().lower()

    def find_col(self, df, keyword, exact=False):
        key = keyword.strip().lower()
        for col in df.columns:
            cname = self.normalize_colname(col)
            if (exact and cname == key) or (not exact and key in cname):
                return col
        return None

    def find_sheet(self, xls, keyword):
        # (注意: xls 现在是 pd.ExcelFile 对象)
        for s in xls.sheet_names:
            if keyword in s:
                return s
        raise ValueError(f"❌ 未找到包含关键词「{keyword}」的sheet")

    def normalize_num(self, val):
        if pd.isna(val): return None
        s = str(val).replace(",", "").strip()
        if s in ["", "-", "nan"]: return None
        try:
            if "%" in s: return float(s.replace("%", "")) / 100
            return float(s)
        except ValueError:
            return s

    def find_file(self, files_list_dict, keyword):
        for kw, f_path in files_list_dict.items():
            if keyword in kw:
                return f_path
        raise FileNotFoundError(f"❌ 未找到包含关键词「{keyword}」的文件")

    def prepare_ref_df(self, ref_df, mapping, prefix):
        contract_col = self.find_col(ref_df, "合同")

        if not contract_col:
            self._log(f"⚠️ 在 {prefix} 参考表中未找到'合同'列，跳过此数据源。", "WARNING")
            return pd.DataFrame(columns=['__KEY__'])

        std_df = pd.DataFrame()
        std_df['__KEY__'] = self.normalize_contract_key(ref_df[contract_col])

        for main_kw, ref_kw in mapping.items():
            exact = (main_kw == "城市经理")
            ref_col_name = self.find_col(ref_df, ref_kw, exact=exact)

            if ref_col_name:
                s_ref_raw = ref_df[ref_col_name]

                if prefix == 'fk' and main_kw == '租赁期限':
                    s_ref_transformed = pd.to_numeric(s_ref_raw, errors='coerce') * 12
                    std_df[f'ref_{prefix}_{main_kw}'] = s_ref_transformed
                else:
                    std_df[f'ref_{prefix}_{main_kw}'] = s_ref_raw
            else:
                self._log(f"⚠️ 在 {prefix} 参考表中未找到列 (main: '{main_kw}', ref: '{ref_kw}')", "WARNING")

        std_df = std_df.drop_duplicates(subset=['__KEY__'], keep='first')
        return std_df

    def compare_series_vec(self, s_main, s_ref, main_kw):
        merge_failed_mask = s_ref.isna()

        main_is_na = pd.isna(s_main) | (s_main.astype(str).str.strip().isin(["", "nan", "None"]))
        ref_is_na = pd.isna(s_ref) | (s_ref.astype(str).str.strip().isin(["", "nan", "None"]))

        both_are_na = main_is_na & ref_is_na

        if any(k in main_kw for k in ["日期", "时间"]):
            d_main = pd.to_datetime(s_main, errors='coerce')
            d_ref = pd.to_datetime(s_ref, errors='coerce')

            valid_dates_mask = d_main.notna() & d_ref.notna()
            date_diff_mask = (d_main.dt.date != d_ref.dt.date)

            errors = valid_dates_mask & date_diff_mask

        else:
            s_main_norm = s_main.apply(self.normalize_num)
            s_ref_norm = s_ref.apply(self.normalize_num)

            main_is_na_norm = pd.isna(s_main_norm) | (s_main_norm.astype(str).str.strip().isin(["", "nan", "None"]))
            ref_is_na_norm = pd.isna(s_ref_norm) | (s_ref_norm.astype(str).str.strip().isin(["", "nan", "None"]))
            both_are_na_norm = main_is_na_norm & ref_is_na_norm

            is_num_main = s_main_norm.apply(lambda x: isinstance(x, (int, float)))
            is_num_ref = s_ref_norm.apply(lambda x: isinstance(x, (int, float)))
            both_are_num = is_num_main & is_num_ref

            errors = pd.Series(False, index=s_main.index)

            if both_are_num.any():
                num_main = s_main_norm[both_are_num].fillna(0)
                num_ref = s_ref_norm[both_are_num].fillna(0)
                diff = (num_main - num_ref).abs()

                if main_kw == "保证金比例":
                    num_errors = (diff > 0.00500001)
                elif "租赁期限" in main_kw:
                    num_errors = (diff >= 1.0)
                else:
                    num_errors = (diff > 1e-6)

                errors.loc[both_are_num] = num_errors

            not_num_mask = ~both_are_num
            if not_num_mask.any():
                str_main = s_main_norm[not_num_mask].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                str_ref = s_ref_norm[not_num_mask].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

                str_errors = (str_main != str_ref)
                errors.loc[not_num_mask] = str_errors

            errors = errors & ~both_are_na_norm

        final_errors = errors & ~both_are_na
        lookup_failure_mask = merge_failed_mask & ~main_is_na
        final_errors = final_errors & ~lookup_failure_mask
        return final_errors

    # =====================================
    # 🧮 单sheet处理函数 (V5: 高效版)
    # =====================================
    # --- (【重构】: "check_one_sheet" -> "process_one_sheet") ---
    # --- (【重构】: 移除了文件读取, 现在接收一个 DataFrame) ---
    def process_one_sheet(self, main_df, sheet_keyword, ref_dfs_std_dict, mappings_all):
        """
        处理一个已经读取的 DataFrame (来自 '月重卡' 的一个 sheet)
        """
        start_time = time.time()

        # (文件读取逻辑已移到 run_audit_logic)

        if main_df.empty:
            self._log(f"⚠️ 「{sheet_keyword}」为空，跳过。", "WARNING")
            return (0, None, 0, set()), []

        contract_col_main = self.find_col(main_df, "合同")
        if not contract_col_main:
            self._log(f"❌ 在「{sheet_keyword}」中未找到合同列。", "ERROR")
            return (0, None, 0, set()), []

        # (保留原有的 TEMP 文件逻辑)
        output_path = f"TEMP__{sheet_keyword}_审核标注版.xlsx"
        try:
            empty_row = pd.DataFrame([[""] * len(main_df.columns)], columns=main_df.columns)
            pd.concat([empty_row, main_df], ignore_index=True).to_excel(output_path, index=False)
            wb = load_workbook(output_path)
            ws = wb.active
        except Exception as e:
            self._log(f"❌ 创建临时文件 {output_path} 失败: {e}", "ERROR")
            return (0, None, 0, set()), []

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        main_df['__ROW_IDX__'] = main_df.index
        main_df['__KEY__'] = self.normalize_contract_key(main_df[contract_col_main])
        contracts_seen = set(main_df['__KEY__'].dropna())

        merged_df = main_df.copy()
        for prefix, std_df in ref_dfs_std_dict.items():
            if not std_df.empty:
                merged_df = pd.merge(merged_df, std_df, on='__KEY__', how='left')

        total_errors = 0
        skip_city_manager = [0]
        errors_locations = set()
        row_has_error = pd.Series(False, index=merged_df.index)

        total_comparisons = sum(len(m[0]) for m in mappings_all.values())
        current_comparison = 0

        self._log(f"\n--- 正在检查: {sheet_keyword} ---", "INFO")
        for prefix, (mapping, std_df) in mappings_all.items():
            if std_df.empty:
                current_comparison += len(mapping)
                continue

            for main_kw, ref_kw in mapping.items():
                current_comparison += 1

                # <--- (修改) 更新状态标签 (进度条已移到外层)
                self._update_status(
                    f"检查「{sheet_keyword}」: {current_comparison}/{total_comparisons} ({prefix} - {main_kw})")

                exact = (main_kw == "城市经理")
                main_col = self.find_col(main_df, main_kw, exact=exact)
                ref_col = f'ref_{prefix}_{main_kw}'

                if not main_col or ref_col not in merged_df.columns:
                    continue

                s_main = merged_df[main_col]
                s_ref = merged_df[ref_col]

                skip_mask = pd.Series(False, index=merged_df.index)
                if main_kw == "城市经理":
                    na_strings = ["", "-", "nan", "none", "null"]
                    skip_mask = pd.isna(s_ref) | s_ref.astype(str).str.strip().isin(na_strings)
                    skip_city_manager[0] += skip_mask.sum()

                errors_mask = self.compare_series_vec(s_main, s_ref, main_kw)
                final_errors_mask = errors_mask & ~skip_mask

                if final_errors_mask.any():
                    total_errors += final_errors_mask.sum()
                    row_has_error |= final_errors_mask

                    bad_indices = merged_df[final_errors_mask]['__ROW_IDX__']
                    for idx in bad_indices:
                        errors_locations.add((idx, main_col))

        self._log(f"  ...「{sheet_keyword}」比对完成，正在生成标注文件...", "INFO")
        self._update_status(f"「{sheet_keyword}」比对完成...")

        original_cols_list = list(main_df.drop(columns=['__ROW_IDX__', '__KEY__']).columns)
        col_name_to_idx = {name: i + 1 for i, name in enumerate(original_cols_list)}

        for (row_idx, col_name) in errors_locations:
            if col_name in col_name_to_idx:
                ws.cell(row_idx + 3, col_name_to_idx[col_name]).fill = red_fill

        if contract_col_main in col_name_to_idx:
            contract_col_excel_idx = col_name_to_idx[contract_col_main]
            error_row_indices = merged_df[row_has_error]['__ROW_IDX__']
            for row_idx in error_row_indices:
                ws.cell(row_idx + 3, contract_col_excel_idx).fill = yellow_fill

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        try:
            os.remove(output_path)
            self._log(f"  ...已清理临时文件: {output_path}", "INFO")
        except OSError as e:
            self._log(f"警告：无法删除临时文件 {output_path}: {e}", "WARNING")

        files_to_save = [
            (f"记录表_{sheet_keyword}_审核标注版.xlsx", output)
        ]
        output_errors_only = None

        if row_has_error.any():
            try:
                df_errors_only = merged_df.loc[row_has_error, original_cols_list].copy()
                original_indices_with_error = merged_df.loc[row_has_error, '__ROW_IDX__']

                original_idx_to_new_excel_row = {
                    original_idx: new_row_num
                    for new_row_num, original_idx in enumerate(original_indices_with_error, start=2)
                }

                wb_errors = Workbook()
                ws_errors = wb_errors.active

                for r in dataframe_to_rows(df_errors_only, index=False, header=True):
                    ws_errors.append(r)

                for (original_row_idx, col_name) in errors_locations:

                    if original_row_idx in original_idx_to_new_excel_row:
                        new_row = original_idx_to_new_excel_row[original_row_idx]

                        if col_name in col_name_to_idx:
                            new_col = col_name_to_idx[col_name]
                            ws_errors.cell(row=new_row, column=new_col).fill = red_fill

                output_errors_only = BytesIO()
                wb_errors.save(output_errors_only)
                output_errors_only.seek(0)

                files_to_save.append(
                    (f"记录表_{sheet_keyword}_仅错误行_标红.xlsx", output_errors_only)
                )
            except Exception as e:
                self._log(f"❌ 生成“仅错误行”文件时出错: {e}", "ERROR")

        elapsed = time.time() - start_time
        self._log(f"✅ {sheet_keyword} 检查完成，共 {total_errors} 处错误，用时 {elapsed:.2f} 秒。", "SUCCESS")

        stats = (total_errors, elapsed, skip_city_manager[0], contracts_seen)
        return stats, files_to_save

    # =====================================
    # 🚀 主执行函数 (V5: 高效模式重构)
    # =====================================
    def run_audit_logic(self):
        """
        这是在后台线程中运行的主函数。
        """
        try:
            self._log("ℹ️ 阶段 1/5: 正在解析文件路径...", "INFO")
            main_file = self.find_file(self.uploaded_files, "月重卡")
            fk_file = self.find_file(self.uploaded_files, "放款明细")
            zd_file = self.find_file(self.uploaded_files, "字段")
            ec_file = self.find_file(self.uploaded_files, "二次明细")

            self._log("ℹ️ 阶段 2/5: 正在读取参考文件...", "INFO")

            self._log("... 正在打开放款明细 (fk_file)", "INFO")
            with pd.ExcelFile(fk_file, engine='calamine') as xls_fk:
                fk_df = pd.read_excel(xls_fk, sheet_name=self.find_sheet(xls_fk, "威田"))

            self._log("... 正在打开字段 (zd_file)", "INFO")
            with pd.ExcelFile(zd_file, engine='calamine') as xls_zd:
                zd_df = pd.read_excel(xls_zd, sheet_name=self.find_sheet(xls_zd, "重卡"))
                contract_col_zd = self.find_col(zd_df, "合同")

            self._log("... 正在打开二次明细 (ec_file)", "INFO")
            ec_df = pd.read_excel(ec_file, engine='calamine')

        except Exception as e:
            self._log(f"❌ 文件读取失败: {e}\n\n请检查文件是否正确，特别是“威田”和“重卡”sheet是否存在。", "ERROR")
            self.root.after(0, messagebox.showerror, "❌ 文件读取失败",
                            f"读取文件或Sheet时出错: {e}\n\n请检查文件是否正确，特别是“威田”和“重卡”sheet是否存在。")
            self._set_gui_state(is_running=False)
            return  # 停止线程

        try:
            self._log("ℹ️ 阶段 3/5: 正在预处理参考数据...", "INFO")
            mapping_fk = {
                "授信方": "授信方",
                "租赁本金": "租赁本金",
                "租赁期限": "租赁期限",
                "挂车台数": "挂车数量",
                "起租收益率": "XIRR"
            }
            mapping_zd = {"保证金比例": "保证金比例_2", "项目提报人": "提报", "起租时间": "起租日_商",
                          "客户经理": "客户经理_资产", "所属省区": "区域", "主车台数": "主车台数",
                          "城市经理": "城市经理"}
            mapping_ec = {"二次时间": "出本流程时间"}

            mappings_all = {
                'fk': (mapping_fk, None),
                'zd': (mapping_zd, None),
                'ec': (mapping_ec, None)
            }

            fk_std = self.prepare_ref_df(fk_df, mapping_fk, 'fk')
            zd_std = self.prepare_ref_df(zd_df, mapping_zd, 'zd')
            ec_std = self.prepare_ref_df(ec_df, mapping_ec, 'ec')

            ref_dfs_std_dict = {'fk': fk_std, 'zd': zd_std, 'ec': ec_std}

            mappings_all['fk'] = (mapping_fk, fk_std)
            mappings_all['zd'] = (mapping_zd, zd_std)
            mappings_all['ec'] = (mapping_ec, ec_std)

            self._log("✅ 参考数据预处理完成。", "SUCCESS")

            sheet_keywords = ["起租", "二次", "部分担保", "随州", "驻店客户"]
            self._log(f"ℹ️ 将审核以下 {len(sheet_keywords)} 个 sheet: {', '.join(sheet_keywords)}", "INFO")

            total_all = elapsed_all = skip_total = 0
            contracts_seen_all_sheets = set()

            self.all_files_to_save = []

            self._log("ℹ️ 阶段 4/5: 正在执行主流程审核...", "INFO")

            # --- (【V5 高效模式重构】) ---
            self._log(f"... 正在打开主文件 '{main_file}' (仅一次)...", "INFO")
            try:
                # 1. 只打开主文件一次
                with pd.ExcelFile(main_file, engine='calamine') as xls_main:

                    # 2. 循环处理 keywords
                    for i, kw in enumerate(sheet_keywords):
                        self._update_progress((i + 1) / len(sheet_keywords))  # <--- 进度条在外层

                        try:
                            # 3. 找到 sheet
                            target_sheet = self.find_sheet(xls_main, kw)

                            # 4. 读取该 sheet
                            self._update_status(f"正在读取 sheet: {target_sheet}...")
                            main_df = pd.read_excel(xls_main, sheet_name=target_sheet, header=1)

                        except ValueError:
                            self._log(f"⚠️ 未找到包含「{kw}」的sheet，跳过。", "WARNING")
                            continue
                        except Exception as e:
                            self._log(f"❌ 读取「{kw}」时出错: {e}", "ERROR")
                            continue

                            # 5. 处理该 sheet (传入 DataFrame)
                        (count, used, skipped, seen), files_list = self.process_one_sheet(
                            main_df,
                            kw,  # (pass the keyword for logging/filenames)
                            ref_dfs_std_dict,
                            mappings_all
                        )

                        # 6. 收集结果
                        self.all_files_to_save.extend(files_list)
                        total_all += count
                        elapsed_all += used or 0
                        skip_total += skipped
                        contracts_seen_all_sheets.update(seen or set())

            except Exception as e:
                self._log(f"❌❌ 无法打开主文件 '月重卡': {e}", "ERROR")
                self._log(traceback.format_exc(), "ERROR")
                raise  # 重新抛出, 被外层 "finally" 捕获

            # --- (高效模式结束) ---

            self._log(f"\n=====================================", "INFO")
            self._log(f"🎯 全部审核完成，共 {total_all} 处错误，总耗时 {elapsed_all:.2f} 秒。", "SUCCESS")
            self._log(f"=====================================", "INFO")

            self._log("ℹ️ 阶段 5/5: 正在执行漏填检查...", "INFO")

            try:
                field_contracts = zd_df[contract_col_zd].dropna().astype(str).str.strip()
                col_car_manager = self.find_col(zd_df, "是否车管家", exact=True)
                col_bonus_type = self.find_col(zd_df, "提成类型", exact=True)
                missing_contracts_mask = (~field_contracts.isin(contracts_seen_all_sheets))

                if col_car_manager:
                    missing_contracts_mask &= ~(zd_df[col_car_manager].astype(str).str.strip().str.lower() == "是")
                if col_bonus_type:
                    missing_contracts_mask &= ~(
                        zd_df[col_bonus_type].astype(str).str.strip().isin(["联合租赁", "驻店"])
                    )

                zd_df_missing = zd_df.copy()
                zd_df_missing["漏填检查"] = ""
                zd_df_missing.loc[missing_contracts_mask, "漏填检查"] = "❗ 漏填"
                漏填合同数 = zd_df_missing["漏填检查"].eq("❗ 漏填").sum()
                self._log(f"🕵️ 漏填检查：共发现 {漏填合同数} 个合同在记录表中未出现（已排除车管家、联合租赁、驻店）", "INFO")

                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                wb = Workbook()
                ws = wb.active
                for r in dataframe_to_rows(zd_df_missing, index=False, header=True):
                    ws.append(r)

                check_col_idx = -1
                for c_idx, cell in enumerate(ws[1], 1):
                    if cell.value == "漏填检查":
                        check_col_idx = c_idx
                        break

                if check_col_idx > 0:
                    for row in ws.iter_rows(min_row=2, min_col=check_col_idx, max_col=check_col_idx):
                        cell = row[0]
                        if cell.value == "❗ 漏填":
                            cell.fill = yellow_fill

                output_all = BytesIO()
                wb.save(output_all)
                output_all.seek(0)
                self.all_files_to_save.append(("字段表_漏填标注版.xlsx", output_all))

                zd_df_only_missing = zd_df_missing[zd_df_missing["漏填检查"] == "❗ 漏填"].copy()
                if not zd_df_only_missing.empty:
                    wb2 = Workbook()
                    ws2 = wb2.active
                    for r in dataframe_to_rows(zd_df_only_missing, index=False, header=True):
                        ws2.append(r)

                    check_col_idx_2 = -1
                    for c_idx, cell in enumerate(ws2[1], 1):
                        if cell.value == "漏填检查":
                            check_col_idx_2 = c_idx
                            break
                    if check_col_idx_2 > 0:
                        for row in ws2.iter_rows(min_row=2, min_col=check_col_idx_2, max_col=check_col_idx_2):
                            if row[0].value == "❗ 漏填":
                                row[0].fill = yellow_fill

                    out2 = BytesIO()
                    wb2.save(out2)
                    out2.seek(0)
                    self.all_files_to_save.append(("字段表_仅漏填.xlsx", out2))

                self._log("✅ 漏填检查文件已生成。", "SUCCESS")

            except Exception as e:
                self._log(f"❌ 漏填检查失败: {e}", "ERROR")
                self._log(traceback.format_exc(), "ERROR")

            self._log("=====================================", "INFO")
            self._log("✅ 所有检查已完成。正在保存文件...", "INFO")

            saved_count = 0
            for filename, buffer in self.all_files_to_save:
                if buffer is None or buffer.getbuffer().nbytes == 0:
                    self._log(f"⚠️ 文件 {filename} 为空，已跳过。", "WARNING")
                    continue
                save_path = os.path.join(self.output_dir, filename)
                try:
                    with open(save_path, 'wb') as f:
                        f.write(buffer.getvalue())
                    self._log(f"  -> 已保存: {save_path}", "SUCCESS")
                    saved_count += 1
                except Exception as e:
                    self._log(f"❌ 保存失败: {filename}。错误: {e}", "ERROR")

            self._log(f"🎉 全部任务执行完毕。共 {saved_count} 个文件已保存到:\n{self.output_dir}", "SUCCESS")

        except Exception as e:
            self._log(f"❌❌❌ 发生未捕获的严重错误: {e}", "ERROR")
            self._log(traceback.format_exc(), "ERROR")
            self.root.after(0, messagebox.showerror, "严重错误", f"发生未捕获的错误: \n{e}")

        finally:
            self._set_gui_state(is_running=False)
            self._update_status("审核完成。")
            self._update_progress(0)

            self._log("--- 退出维护：正在清理当前临时文件... ---", "INFO")
            temp_files = [f for f in os.listdir() if f.startswith("TEMP__") and f.endswith(".xlsx")]
            if not temp_files:
                self._log("  > 无临时文件需要清理。", "INFO")
            for f in temp_files:
                try:
                    os.remove(f)
                    self._log(f"  > 已清理: {f}", "INFO")
                except OSError as e:
                    self._log(f"  > ⚠️ 清理失败 (可能被占用): {f}. 错误: {e}", "WARNING")

            self.root.after(0, messagebox.showinfo, "任务完成",
                            f"审核已全部完成！\n共 {len(self.all_files_to_save)} 个报告文件已保存到:\n{self.output_dir}")


# =====================================
# 🚀 启动器
# =====================================
if __name__ == "__main__":

    clear_local_cache_on_startup()  # <-- 启动时清理

    try:
        root = tk.Tk()
        app = AuditApp(root)
        root.mainloop()
    except Exception as e:
        print(f"❌ 发生未处理的致命错误: {e}")
        print(traceback.format_exc())
        try:
            tk_error_root = tk.Tk()
            tk_error_root.withdraw()
            messagebox.showerror("❌ 致命错误", f"程序遇到无法处理的错误，即将退出。\n\n错误详情: {e}")
        except:
            pass
    finally:
        print("程序已退出。")