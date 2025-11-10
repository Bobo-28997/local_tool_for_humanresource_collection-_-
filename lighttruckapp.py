# =====================================
# Streamlit App (V2) -> 转换为 Tkinter 桌面应用
#
# 【Calamine 整合版】
# - [新] 所有 pd.read_excel/pd.ExcelFile 调用均已切换到 'calamine' 引擎
# - 包含 Streamlit -> Tkinter 的转换 (UI, 线程等)
# - 包含 pd.ExcelFile 的 'with' 语句修复，防止 "一次性exe" 文件锁定问题
# =====================================

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import unicodedata, re
import time
import threading
import os
import traceback  # 用于更详细的错误日志


# =====================================
# 🛠️ 未改变的工具函数 (来自 Streamlit 脚本)
# =====================================
# (这些函数是纯 Python/Pandas 逻辑, 无需修改)

def normalize_colname(c):
    return str(c).strip().lower()


def find_col(df, keyword, exact=False):
    if df is None:
        return None
    key = keyword.strip().lower()
    for col in df.columns:
        cname = normalize_colname(col)
        if (exact and cname == key) or (not exact and key in cname):
            return col
    return None


def normalize_num(val):
    if pd.isna(val):
        return None
    s = str(val).replace(",", "").strip()
    if s in ["", "-", "nan"]:
        return None
    try:
        if "%" in s:
            s = s.replace("%", "")
            return float(s) / 100
        return float(s)
    except:
        return s


def normalize_text(val):
    if pd.isna(val):
        return ""
    s = str(val)
    s = re.sub(r'[\n\r\t ]+', '', s)
    s = s.replace('\u3000', '')
    s = ''.join(unicodedata.normalize('NFKC', ch) for ch in s)
    return s.lower().strip()


def detect_header_row(file, sheet_name):
    # <--- TKINTER MODIFICATION: 调整为接受文件路径(path)
    try:
        # --- (【Calamine 修改】) ---
        with pd.ExcelFile(file, engine='calamine') as xls:
            # (read_excel 会自动继承 calamine 引擎)
            preview = pd.read_excel(xls, sheet_name=sheet_name, nrows=2, header=None)

        first_row = preview.iloc[0]
        total_cells = len(first_row)
        empty_like = sum(
            (pd.isna(x) or str(x).startswith("Unnamed") or str(x).strip() == "")
            for x in first_row
        )
        empty_ratio = empty_like / total_cells if total_cells > 0 else 0
        return 1 if empty_ratio >= 0.7 else 0
    except Exception as e:
        print(f"Error detecting header for {sheet_name}: {e}")
        return 0  # 默认


def get_header_row(file, sheet_name):
    if any(k in sheet_name for k in ["起租", "二次"]):
        return 1
    return detect_header_row(file, sheet_name)


def normalize_contract_key(series: pd.Series) -> pd.Series:
    s = series.astype(str)
    s = s.str.replace(r"\.0$", "", regex=True)
    s = s.str.strip()
    s = s.str.upper()
    s = s.str.replace('－', '-', regex=False)
    return s


def compare_series_vec(s_main, s_ref, compare_type='text', tolerance=0, multiplier=1):
    """
    (V3: 增加 'num_term' 类型)
    """
    merge_failed_mask = s_ref.isna()
    main_is_na = pd.isna(s_main) | (s_main.astype(str).str.strip().isin(["", "nan", "None"]))
    ref_is_na = pd.isna(s_ref) | (s_ref.astype(str).str.strip().isin(["", "nan", "None"]))
    both_are_na = main_is_na & ref_is_na

    errors = pd.Series(False, index=s_main.index)

    # 2. 日期比较
    if compare_type == 'date':
        d_main = pd.to_datetime(s_main, errors='coerce')
        d_ref = pd.to_datetime(s_ref, errors='coerce')

        valid_dates_mask = d_main.notna() & d_ref.notna()
        date_diff_mask = (d_main.dt.date != d_ref.dt.date)
        errors = valid_dates_mask & date_diff_mask

        one_is_date_one_is_not = (d_main.notna() & d_ref.isna() & ~ref_is_na) | \
                                 (d_main.isna() & ~main_is_na & d_ref.notna())
        errors |= one_is_date_one_is_not

    # 3. 数值比较
    elif compare_type in ['num', 'num_term']:
        s_main_norm = s_main.apply(normalize_num)
        s_ref_norm = s_ref.apply(normalize_num)

        is_num_main = s_main_norm.apply(lambda x: isinstance(x, (int, float)))
        is_num_ref = s_ref_norm.apply(lambda x: isinstance(x, (int, float)))
        both_are_num = is_num_main & is_num_ref

        if both_are_num.any():
            diff = (s_main_norm[both_are_num] - s_ref_norm[both_are_num]).abs()

            if compare_type == 'num_term':
                errors.loc[both_are_num] = (diff >= 1.0)
            else:
                errors.loc[both_are_num] = (diff > (tolerance + 1e-6))

        one_is_num_one_is_not = (is_num_main & ~is_num_ref & ~ref_is_na) | \
                                (~is_num_main & ~main_is_na & is_num_ref)
        errors |= one_is_num_one_is_not

    # 4. 文本比较
    else:  # compare_type == 'text'
        s_main_norm_text = s_main.apply(normalize_text)
        s_ref_norm_text = s_ref.apply(normalize_text)
        errors = (s_main_norm_text != s_ref_norm_text)

    # 5. 最终错误逻辑
    final_errors = errors & ~both_are_na
    lookup_failure_mask = merge_failed_mask & ~main_is_na
    final_errors = final_errors & ~lookup_failure_mask

    return final_errors


# =====================================
# 🖥️ Tkinter 应用主类
# =====================================

class AuditApp:
    def __init__(self, root):
        self.root = root
        # --- (【Calamine 修改】) ---
        self.root.title("📊 人事薪资表自动审核系统-2 (轻卡)")
        self.root.geometry("800x700")

        self.style = ttk.Style()
        self.style.theme_use('clam')  # 'clam', 'alt', 'default', 'classic'

        # --- 实例变量 ---
        self.uploaded_files = {}  # 存储文件路径, e.g. {'项目提成': 'path/to/file.xlsx', ...}
        self.output_dir = ""
        self.required_files = ["项目提成", "放款明细", "二次明细", "产品台账"]

        # --- GUI 布局 ---
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- 1. 文件选择区 ---
        input_frame = ttk.LabelFrame(main_frame, text="1. 输入文件", padding="10")
        input_frame.pack(fill=tk.X, expand=False)

        self.select_files_button = ttk.Button(input_frame, text="选择 4 个 Excel 文件", command=self.select_files)
        self.select_files_button.pack(side=tk.LEFT, padx=(0, 10))

        self.file_status_label = ttk.Label(input_frame, text="尚未选择文件")
        self.file_status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- 2. 输出文件夹选择区 ---
        output_frame = ttk.LabelFrame(main_frame, text="2. 输出文件夹", padding="10")
        output_frame.pack(fill=tk.X, expand=False, pady=5)

        self.select_output_dir_button = ttk.Button(output_frame, text="选择报告保存位置",
                                                   command=self.select_output_dir)
        self.select_output_dir_button.pack(side=tk.LEFT, padx=(0, 10))

        self.output_dir_label = ttk.Label(output_frame, text="尚未选择文件夹")
        self.output_dir_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- 3. 执行区 ---
        run_frame = ttk.Frame(main_frame, padding="10")
        run_frame.pack(fill=tk.X, expand=False)

        self.run_button = ttk.Button(run_frame, text="🚀 开始审核", command=self.run_audit_thread, state=tk.DISABLED)
        self.run_button.pack(fill=tk.X, expand=True)

        # --- 4. 状态和进度区 ---
        status_frame = ttk.LabelFrame(main_frame, text="3. 审核状态", padding="10")
        status_frame.pack(fill=tk.X, expand=False, pady=5)

        self.status_label = ttk.Label(status_frame, text="等待开始...")
        self.status_label.pack(fill=tk.X, expand=True)

        self.progress_bar = ttk.Progressbar(status_frame, orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress_bar.pack(fill=tk.X, expand=True, pady=5)

        # --- 5. 日志区 ---
        log_frame = ttk.LabelFrame(main_frame, text="4. 运行日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.log_widget = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=20, state=tk.DISABLED)
        self.log_widget.pack(fill=tk.BOTH, expand=True)

        # 定义日志颜色
        self.log_widget.tag_config('INFO', foreground='black')
        self.log_widget.tag_config('SUCCESS', foreground='green')
        self.log_widget.tag_config('WARNING', foreground='orange')
        self.log_widget.tag_config('ERROR', foreground='red', font=('Helvetica', '9', 'bold'))

    # =====================================
    # 🖥️ GUI 交互方法
    # =====================================

    def _log(self, message, level='INFO'):
        """线程安全的日志更新"""
        self.root.after(0, self.update_log_widget, message, level)

    def update_log_widget(self, message, level):
        self.log_widget.config(state=tk.NORMAL)
        self.log_widget.insert(tk.END, f"{message}\n", level.upper())
        self.log_widget.see(tk.END)  # 自动滚动到底部
        self.log_widget.config(state=tk.DISABLED)

    def _update_status(self, text):
        """线程安全的状态标签更新"""
        self.root.after(0, self.status_label.config, {'text': text})

    def _update_progress(self, value):
        """线程安全的进度条更新 (value: 0.0 to 1.0)"""
        self.root.after(0, self.progress_bar.config, {'value': value * 100})

    def _set_gui_state(self, is_running):
        """线程安全地切换按钮状态"""
        self.root.after(0, self.toggle_buttons, is_running)

    def toggle_buttons(self, is_running):
        state = tk.DISABLED if is_running else tk.NORMAL
        self.select_files_button.config(state=state)
        self.select_output_dir_button.config(state=state)
        if not is_running and self.check_ready(silent=True):
            self.run_button.config(state=tk.NORMAL)
        else:
            self.run_button.config(state=tk.DISABLED)

    def check_ready(self, silent=False):
        """检查所有条件是否满足，以启用“开始审核”按钮"""
        ready = len(self.uploaded_files) == len(self.required_files) and self.output_dir
        if ready:
            self.run_button.config(state=tk.NORMAL)
            if not silent:
                self._log("✅ 所有文件和输出目录已就绪，可以开始审核。", "SUCCESS")
        else:
            self.run_button.config(state=tk.DISABLED)
        return ready

    def select_files(self):
        self.uploaded_files = {}  # 重置

        file_paths = filedialog.askopenfilenames(
            title=f"请选择所有 {len(self.required_files)} 个文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if not file_paths:
            self.file_status_label.config(text="未选择文件")
            self.check_ready()
            return

        found_count = 0
        missing_files = list(self.required_files)  # 复制一份

        for keyword in self.required_files:
            found_for_keyword = False
            for path in file_paths:
                filename = os.path.basename(path)
                if keyword in filename:
                    self.uploaded_files[keyword] = path
                    found_count += 1
                    if keyword in missing_files:
                        missing_files.remove(keyword)
                    found_for_keyword = True
                    break  # 一个关键字只匹配一个文件

        self.file_status_label.config(
            text=f"已选择 {found_count} / {len(self.required_files)} 个文件。"
        )

        if missing_files:
            self._log(f"⚠️ 仍缺少文件: {', '.join(missing_files)}", "WARNING")

        self.check_ready()

    def select_output_dir(self):
        dir_path = filedialog.askdirectory(title="选择报告保存的文件夹")
        if dir_path:
            self.output_dir = dir_path
            self.output_dir_label.config(text=f"将保存到: {dir_path}")
        else:
            self.output_dir = ""
            self.output_dir_label.config(text="尚未选择文件夹")
        self.check_ready()

    def run_audit_thread(self):
        """“开始审核”按钮的入口，启动一个新线程"""
        self._set_gui_state(is_running=True)
        self.log_widget.config(state=tk.NORMAL)
        self.log_widget.delete('1.0', tk.END)  # 清空日志
        self.log_widget.config(state=tk.DISABLED)

        self._log("🚀 审核任务开始...", "INFO")

        # 启动后台线程
        threading.Thread(target=self.run_audit_logic, daemon=True).start()

    # =====================================
    # 🧮 核心审核逻辑 (移植自 Streamlit)
    # =====================================

    def prepare_one_ref_df(self, ref_df, ref_contract_col, required_cols, prefix):
        if ref_df is None:
            self._log(f"⚠️ 参考文件 '{prefix}' 未加载 (df is None)。", "WARNING")
            return pd.DataFrame(columns=['__KEY__'])

        if ref_contract_col is None:
            self._log(f"⚠️ 在 {prefix} 参考表中未找到'合同'列，跳过此数据源。", "WARNING")
            return pd.DataFrame(columns=['__KEY__'])

        cols_to_extract = []
        col_mapping = {}

        for col_kw in required_cols:
            actual_col = find_col(ref_df, col_kw)

            if actual_col:
                cols_to_extract.append(actual_col)
                col_mapping[actual_col] = f"ref_{prefix}_{col_kw}"
            else:
                self._log(f"⚠️ 在 {prefix} 参考表中未找到列 (关键字: '{col_kw}')", "WARNING")

        if not cols_to_extract:
            self._log(f"⚠️ 在 {prefix} 参考表中未找到任何所需字段，跳过。", "WARNING")
            return pd.DataFrame(columns=['__KEY__'])

        cols_to_extract.append(ref_contract_col)
        cols_to_extract_unique = list(set(cols_to_extract))
        valid_cols = [col for col in cols_to_extract_unique if col in ref_df.columns]
        std_df = ref_df[valid_cols].copy()

        std_df['__KEY__'] = normalize_contract_key(std_df[ref_contract_col])
        std_df = std_df.rename(columns=col_mapping)
        final_cols = ['__KEY__'] + list(col_mapping.values())
        final_cols_in_df = [col for col in final_cols if col in std_df.columns]
        std_df = std_df[final_cols_in_df]
        std_df = std_df.drop_duplicates(subset=['__KEY__'], keep='first')
        return std_df

    def audit_sheet_vec(self, sheet_name, xls_main, main_file_path, all_std_dfs, mapping_rules_vec):

        try:
            # (注意: get_header_row 仍然需要 main_file_path 来读取预览)
            # (它内部的 pd.ExcelFile 已被修改为 calamine)
            header_offset = get_header_row(main_file_path, sheet_name)

            # (xls_main 已使用 calamine 打开, 此处自动继承)
            main_df = pd.read_excel(xls_main, sheet_name=sheet_name, header=header_offset)

            self._log(f"📘 审核中：{sheet_name}（header={header_offset}）", "INFO")

            contract_col_main = find_col(main_df, "合同")
            if not contract_col_main:
                self._log(f"❌ {sheet_name} 中未找到“合同”列，已跳过。", "ERROR")
                return None, 0

            main_df['__ROW_IDX__'] = main_df.index
            main_df['__KEY__'] = normalize_contract_key(main_df[contract_col_main])

            merged_df = main_df.copy()
            for std_df in all_std_dfs.values():
                if not std_df.empty:
                    merged_df = pd.merge(merged_df, std_df, on='__KEY__', how='left')

            total_errors = 0
            errors_locations = set()
            row_has_error = pd.Series(False, index=merged_df.index)

            self._update_progress(0)

            total_comparisons = len(mapping_rules_vec)
            current_comparison = 0

            for main_kw, comparisons in mapping_rules_vec.items():
                current_comparison += 1

                main_col = find_col(main_df, main_kw)
                if not main_col:
                    continue

                self._update_status(f"检查「{sheet_name}」: {main_kw}...")

                field_error_mask = pd.Series(False, index=merged_df.index)

                for (ref_col, compare_type, tol, mult) in comparisons:
                    if ref_col not in merged_df.columns:
                        continue

                    s_main = merged_df[main_col]
                    s_ref = merged_df[ref_col]

                    skip_mask = pd.Series(False, index=merged_df.index)

                    if main_kw == "城市经理":
                        na_mask = pd.isna(s_ref)
                        str_val = s_ref.astype(str).str.strip().str.lower()
                        str_mask = str_val.isin(["", "nan", "none", "null", "0", "0.0"])
                        skip_mask = na_mask | str_mask

                    errors_mask = compare_series_vec(s_main, s_ref, compare_type, tol, mult)
                    final_errors_mask = errors_mask & ~skip_mask
                    field_error_mask |= final_errors_mask

                if field_error_mask.any():
                    total_errors += field_error_mask.sum()
                    row_has_error |= field_error_mask

                    bad_indices = merged_df[field_error_mask]['__ROW_IDX__']
                    for idx in bad_indices:
                        errors_locations.add((idx, main_col))

                self._update_progress(current_comparison / total_comparisons)

            self._update_status(f"「{sheet_name}」比对完成，正在生成标注文件...")

            # 5. === 快速写入 Excel 并标注 ===
            wb = Workbook()
            ws = wb.active
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            original_cols_list = list(main_df.drop(columns=['__ROW_IDX__', '__KEY__']).columns)
            col_name_to_idx = {name: i + 1 for i, name in enumerate(original_cols_list)}

            if header_offset > 0:
                for _ in range(header_offset):
                    ws.append([""] * len(original_cols_list))

            for r in dataframe_to_rows(main_df[original_cols_list], index=False, header=True):
                ws.append(r)

            for (row_idx, col_name) in errors_locations:
                if col_name in col_name_to_idx:
                    excel_row = row_idx + 1 + header_offset + 1
                    excel_col = col_name_to_idx[col_name]
                    ws.cell(excel_row, excel_col).fill = red_fill

            if contract_col_main in col_name_to_idx:
                contract_col_excel_idx = col_name_to_idx[contract_col_main]
                error_row_indices = merged_df[row_has_error]['__ROW_IDX__']
                for row_idx in error_row_indices:
                    excel_row = row_idx + 1 + header_offset + 1
                    ws.cell(excel_row, contract_col_excel_idx).fill = yellow_fill

            # 6. (修改) 导出到文件
            save_path_full = os.path.join(self.output_dir, f"{sheet_name}_审核标注版.xlsx")
            wb.save(save_path_full)
            self._log(f"📥 报告已保存: {save_path_full}", "SUCCESS")

            # 7. (修改) 导出仅含错误行的文件 (带标红)
            if row_has_error.any():
                try:
                    df_errors_only = merged_df.loc[row_has_error, original_cols_list].copy()
                    original_indices_with_error = merged_df.loc[row_has_error, '__ROW_IDX__']
                    original_idx_to_new_excel_row = {
                        original_idx: new_row_num
                        for new_row_num, original_idx in enumerate(original_indices_with_error, start=2)
                    }
                    wb_errors = Workbook()
                    ws_errors = wb_errors.active
                    for r in dataframe_to_rows(df_errors_only, index=False, header=True):
                        ws_errors.append(r)
                    for (original_row_idx, col_name) in errors_locations:
                        if original_row_idx in original_idx_to_new_excel_row:
                            new_row = original_idx_to_new_excel_row[original_row_idx]
                            if col_name in col_name_to_idx:
                                new_col = col_name_to_idx[col_name]
                                ws_errors.cell(row=new_row, column=new_col).fill = red_fill

                    save_path_errors = os.path.join(self.output_dir, f"{sheet_name}_仅错误行_标红.xlsx")
                    wb_errors.save(save_path_errors)
                    self._log(f"📥 仅错误行报告已保存: {save_path_errors}", "SUCCESS")

                except Exception as e:
                    self._log(f"❌ 生成“仅错误行”文件时出错: {e}", "ERROR")

            self._log(f"✅ {sheet_name} 审核完成，共发现 {total_errors} 处错误", "SUCCESS")

            return main_df.drop(columns=['__ROW_IDX__', '__KEY__']), total_errors

        except Exception as e:
            self._log(f"❌❌❌ 在处理 {sheet_name} 时发生严重错误: {e}", "ERROR")
            self._log(traceback.format_exc(), "ERROR")
            return None, 0

    def run_audit_logic(self):

        try:
            # =====================================
            # 🛠️ 文件路径准备
            # =====================================

            main_file = self.uploaded_files.get("项目提成")
            ec_file = self.uploaded_files.get("二次明细")
            fk_file = self.uploaded_files.get("放款明细")
            product_file = self.uploaded_files.get("产品台账")

            if not all([main_file, ec_file, fk_file, product_file]):
                self._log("❌ 内部错误：文件字典不完整。", "ERROR")
                return

            self._log("ℹ️ -------------------------------", "INFO")
            self._log("ℹ️ 阶段 1/3: 正在读取并预处理参考文件...", "INFO")

            # --- (【Calamine 修改】) ---
            ec_df = pd.read_excel(ec_file, engine='calamine')
            product_df = pd.read_excel(product_file, engine='calamine')

            commission_df = None
            all_std_dfs = {}
            contract_col_comm = None

            # --- (【Calamine 修改】) ---
            with pd.ExcelFile(fk_file, engine='calamine') as fk_xls:
                commission_sheets = [s for s in fk_xls.sheet_names if "提成" in s]
                if not commission_sheets:
                    self._log("❌ 在 '放款明细' 文件中未找到任何包含 '提成' 的sheet！程序已停止。", "ERROR")
                    return

                self._log(f"ℹ️ 正在从 '放款明细' 加载 {len(commission_sheets)} 个 '提成' sheet...", "INFO")
                # (read_excel 会自动继承 calamine 引擎)
                commission_df_list = [pd.read_excel(fk_xls, sheet_name=s) for s in commission_sheets]
                fk_commission_df = pd.concat(commission_df_list, ignore_index=True)

                fk_df = fk_commission_df
                commission_df = fk_commission_df

                contract_col_ec = find_col(ec_df, "合同")
                contract_col_fk = find_col(fk_df, "合同")
                contract_col_comm = find_col(commission_df, "合同")
                contract_col_product = find_col(product_df, "合同")

                mapping_rules_vec = {
                    "起租日期": [("ref_ec_起租日_商", 'date', 0, 1)],
                    "租赁本金": [("ref_fk_租赁本金", 'num', 0, 1)],
                    "收益率": [("ref_fk_XIRR", 'num', 0.005, 1)],
                    "操作人": [("ref_fk_提报人员", 'text', 0, 1)],
                    "客户经理": [("ref_fk_提报人员", 'text', 0, 1)],
                    "城市经理": [("ref_fk_城市经理", 'text', 0, 1)],
                    "完成二次交接时间": [("ref_ec_出本流程时间", 'date', 0, 1)],
                    "年化MIN": [("ref_product_年化", 'num', 0.005, 1)],
                    "年限": [("ref_fk_租赁期限", 'num_term', 0, 0)]
                }

                ec_cols = ["起租日_商", "出本流程时间"]
                fk_cols = ["租赁本金", "提报人员", "城市经理", "租赁期限", "XIRR"]
                product_cols = ["年化"]

                ec_std = self.prepare_one_ref_df(ec_df, contract_col_ec, ec_cols, "ec")
                fk_std = self.prepare_one_ref_df(fk_df, contract_col_fk, fk_cols, "fk")
                product_std = self.prepare_one_ref_df(product_df, contract_col_product, product_cols, "product")

                all_std_dfs = {"ec": ec_std, "fk": fk_std, "product": product_std}

            self._log("✅ 参考文件预处理完成。", "SUCCESS")

            # =====================================
            # 🛠️ (修改) 阶段 2/3: 执行主流程
            # =====================================
            self._log("ℹ️ -------------------------------", "INFO")
            self._log("ℹ️ 阶段 2/3: 正在执行主流程审核...", "INFO")

            all_contracts_in_sheets = set()

            # --- (【Calamine 修改】) ---
            with pd.ExcelFile(main_file, engine='calamine') as xls_main:
                target_sheets = [
                    s for s in xls_main.sheet_names
                    if any(k in s for k in ["起租", "二次", "平台工", "独立架构", "低价值"])
                ]

                if not target_sheets:
                    self._log("⚠️ 未在 '项目提成' 文件中找到任何目标 sheet。", "WARNING")
                else:
                    for sheet_name in target_sheets:
                        df, _ = self.audit_sheet_vec(
                            sheet_name,
                            xls_main,
                            main_file,
                            all_std_dfs,
                            mapping_rules_vec
                        )

                        if df is not None:
                            col = find_col(df, "合同")
                            if col:
                                normalized_contracts = normalize_contract_key(df[col].dropna())
                                all_contracts_in_sheets.update(normalized_contracts)

            self._log("✅ 主流程审核完成。", "SUCCESS")

            # =====================================
            # 🛠️ 阶段 3/3: 漏填检测
            # =====================================
            self._log("ℹ️ -------------------------------", "INFO")
            self._log("ℹ️ 阶段 3/3: 正在执行漏填检测...", "INFO")

            if commission_df is not None and contract_col_comm:
                commission_contracts = set(normalize_contract_key(commission_df[contract_col_comm].dropna()))

                missing_contracts = sorted(list(commission_contracts - all_contracts_in_sheets))

                self._log(f"📋 共 {len(missing_contracts)} 个合同在六张表中未出现。", "INFO")

                if missing_contracts:
                    wb_miss = Workbook()
                    ws_miss = wb_miss.active
                    ws_miss.cell(1, 1, "未出现在任一表中的合同号")
                    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    for i, cno in enumerate(missing_contracts, start=2):
                        ws_miss.cell(i, 1, cno).fill = yellow

                    save_path_missing = os.path.join(self.output_dir, "漏填合同号列表.xlsx")
                    wb_miss.save(save_path_missing)
                    self._log(f"📥 漏填合同列表已保存: {save_path_missing}", "SUCCESS")
                else:
                    self._log("✅ 所有提成sheet合同号均已出现在六张表中，无漏填。", "SUCCESS")

            else:
                self._log("⚠️ 跳过漏填检测，因为 'commission_df' 未被成功加载。", "WARNING")

            self._log("🎉🎉🎉 所有审核任务已完成！", "SUCCESS")

        except Exception as e:
            self._log(f"❌❌❌ 发生未捕获的严重错误: {e}", "ERROR")
            self._log(traceback.format_exc(), "ERROR")
            self.root.after(0, messagebox.showerror, "严重错误", f"发生未捕获的错误: \n{e}")

        finally:
            self._set_gui_state(is_running=False)
            self._update_status("审核完成。")
            self._update_progress(0)
            self.root.after(0, messagebox.showinfo, "任务完成", "审核已全部完成！")


# =====================================
# 🚀 启动应用
# =====================================
if __name__ == "__main__":
    root = tk.Tk()
    app = AuditApp(root)
    root.mainloop()