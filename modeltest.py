import json
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
import os

# --- 1. 定义您的核心Prompt (从您的问题中复制) ---
# 确保这个Prompt与您在LM Studio中测试时完全一致
SYSTEM_PROMPT = """你是一个极其严谨、注重细节的法律文书提取机器人。
你的唯一任务是：严格按照“JSON Output”的格式，从“法律文书原文”中提取信息。
--- 关键规则 ---
1. 你必须严格模仿我给出的示例1。
2. 【最重要的陷阱】：`case_number` (案号) 是最难的字段。它常常【没有'案号：'这样的标签】，而是作为单独一行【出现在文书的最顶部】。你的首要任务就是在文书的开头几行里找到那个格式类似 '(xxxx) ...号' 的字符串，并将其识别为 `case_number`。
3. 【地址陷阱】：`address`字段必须严格提取在 `住所地：` 或 `住 ` 之后、逗号之前的文本。
4. 【多重字段陷阱】：如果有多个“被告”，请将他们的信息（如 name, address）用 `、` (顿号) 连接合并，如范例所示。
5. 模仿示例中 `null` 的用法。如果找不到字段，必须返回 `null`。
--- 输出格式铁则 (最重要) ---
6. 你的回答必须，也只能是一个JSON对象。
7. 你的回答必须严格以 `{` (左大括号) 开始。
8. 你的回答必须严格以 `}` (右大括号) 结束。
9. 在 `{...}` 之外，绝对不允许有任何其他文本、注释、Markdown标记 (如 ```json 或 ```) 或任何形式的问候语/解释。
--- 示例1 (黄金范例) --- 
法律文书原文:
(2025) 京01民初789号
北京市海淀区人民法院 民 事 判 决 书
原告：北京A科技公司，住所地：北京市海淀区中关村大街1号，统一社会信用代码：12344usb223a。
法定代表人：王明，董事长。
委托诉讼代理人：李杨，北京基米律师事务所律师。
被告：上海B数据服务有限公司，住所地：北京市海淀区中关村大街9号，统一社会信用代码：12344iws223a。
被告：牛二，住北京市海淀区中关村大街3号，身份证号：12392839804012。
本院经审理查明...原告因本案诉讼支出代理费5000元。...（此处省略正文）
判决如下： 一、被告牛二，上海B数据服务有限公司应于本判决生效之日起十日内，共同支付原告北京A科技公司合同款100,000元，原告代理费 5 000 元，共计 105,000元，于本判决生效后十日内履行完毕。 二、驳回原告其他诉讼请求。
案件受理费1000元，由被告负担。
审判员：张伟 二〇二五年十月一日 （本判决书为虚构）
JSON Output: { "type": "民事判决书", "case_number": "(2025) 京01民初789号",
"court_name": "北京市海淀区人民法院",
"judgment_date": "二〇二五年十月一日",
"plaintiff": { "name": "北京A科技公司", "address": "北京市海淀区中关村大街1号", "legal_rep": "王明", "authorized_agent": "李杨，北京基米律师事务所律师", "agent_price": "5000元"
}, "defendant": { "name": "上海B数据服务有限公司、牛二", "address": "北京市海淀区中关村大街9号、北京市海淀区中关村大街3号",
"legal_rep": null }, "judge": { "judge_main":
"张伟", "judge_execution": null },
"judgment_summary": "被告牛二、上海B数据服务有限公司应于本判决生效之日起十日内，支付原告北京A科技公司合同款100,000元，原告代理费 5 000 元，共计 105,000元，于本判决生效后十日内履行完毕。" }
--- 真实任务 ---
现在，请【严格遵守上述所有规则，尤其是关于'案号'、'地址'和'多重字段'的陷阱规则】，处理我提供的文件。"""

# --- 2. 定义文件和API路径 ---
LM_STUDIO_API_URL = "http://localhost:1234/v1/chat/completions"
INPUT_TEXT_FILE = "sample.txt"
OUTPUT_EXCEL_FILE = "template.xlsx"


def call_lm_studio(raw_text):
    """调用LM Studio API并获取结构化的JSON输出"""
    headers = {"Content-Type": "application/json"}

    # 构建发送给API的 "messages"
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": f"法律文书原文:\n{raw_text}"}
    ]

    # 构建请求体
    payload = {
        "messages": messages,
        "temperature": 0.0,  # 结构化提取，温度设为0，确保严谨
        "max_tokens": 2048,
        "stream": False
    }

    print("正在向 LM Studio 发送请求，请稍候...")

    try:
        response = requests.post(LM_STUDIO_API_URL, headers=headers, json=payload, timeout=300)
        response.raise_for_status()  # 如果请求失败（如4xx或5xx），则引发异常

        # 解析返回的JSON
        api_response = response.json()

        # 提取模型生成的JSON *字符串*
        model_output_string = api_response['choices'][0]['message']['content']

        print("--- 模型原始输出 (JSON 字符串) ---")
        print(model_output_string)
        print("---------------------------------")

        # 将这个JSON字符串解析为Python字典
        extracted_data = json.loads(model_output_string)
        return extracted_data

    except requests.exceptions.ConnectionError:
        print(f"错误：无法连接到 {LM_STUDIO_API_URL}。")
        print("请确保 LM Studio 服务器正在运行。")
        return None
    except requests.exceptions.Timeout:
        print("错误：请求超时。模型可能需要更长时间处理。")
        return None
    except requests.exceptions.RequestException as e:
        print(f"错误：API请求失败: {e}")
        return None
    except json.JSONDecodeError:
        print("错误：模型未返回有效的JSON。请检查模型输出或Prompt。")
        print(f"接收到的内容: {model_output_string}")
        return None
    except KeyError:
        print("错误：API返回的结构不符合预期。")
        print(f"API原始响应: {api_response}")
        return None


def write_to_excel(data, excel_path):
    """将数据写入指定的Excel文件"""

    # 1. 扁平化数据 (将嵌套的JSON转换为平坦的字典)
    # 使用 .get() 来安全地访问，如果键不存在则返回 None 或 {}
    plaintiff_data = data.get('plaintiff', {})
    defendant_data = data.get('defendant', {})
    judge_data = data.get('judge', {})

    flat_data = {
        'type': data.get('type'),
        'case_number': data.get('case_number'),
        'court_name': data.get('court_name'),
        'judgment_date': data.get('judgment_date'),
        'plaintiff_name': plaintiff_data.get('name'),
        'plaintiff_address': plaintiff_data.get('address'),
        'plaintiff_legal_rep': plaintiff_data.get('legal_rep'),
        'plaintiff_authorized_agent': plaintiff_data.get('authorized_agent'),
        'plaintiff_agent_price': plaintiff_data.get('agent_price'),
        'defendant_name': defendant_data.get('name'),
        'defendant_address': defendant_data.get('address'),
        'defendant_legal_rep': defendant_data.get('legal_rep'),
        'judge_main': judge_data.get('judge_main'),
        # 'judge_execution': judge_data.get('judge_execution'), # 如果需要，取消注释
        'judgment_summary': data.get('judgment_summary')
    }

    print(f"正在准备将以下数据写入Excel: \n{json.dumps(flat_data, indent=2, ensure_ascii=False)}")

    # 2. 加载工作簿和工作表
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        print(f"成功加载Excel文件: {excel_path}")
    except InvalidFileException:
        print(f"错误：{excel_path} 不是有效的Excel文件或已损坏。")
        return
    except FileNotFoundError:
        print(f"错误：找不到 {excel_path}。")
        print("请确保Excel文件与脚本在同一目录下，并且名称正确。")
        return

    # 3. 读取表头 (第一行)
    headers = [cell.value for cell in ws[1]]
    if not headers:
        print("错误：Excel文件 {excel_path} 的第一行没有表头。")
        return

    # 4. 找到下一个空行
    next_row = ws.max_row + 1

    # 5. 写入数据
    # 遍历表头，根据表头名称从 flat_data 中查找数据并填入
    for col_idx, header_name in enumerate(headers, start=1):
        if header_name in flat_data:
            cell_value = flat_data[header_name]
            # 确保 None 值被写入为空单元格，而不是字符串 "None"
            if cell_value is None:
                cell_value = ""
            ws.cell(row=next_row, column=col_idx).value = cell_value
        else:
            # 如果表头在数据中找不到，留空并在控制台提示
            print(f"警告：在数据中未找到表头 '{header_name}'，Excel中对应单元格将留空。")

    # 6. 保存工作簿
    try:
        wb.save(excel_path)
        print(f"成功！数据已追加到 {excel_path} 的第 {next_row} 行。")
    except PermissionError:
        print(f"错误：保存失败。请关闭正在打开的 {excel_path} 文件后再试。")
    except Exception as e:
        print(f"保存Excel时发生未知错误: {e}")


def main():
    """主执行函数"""
    # 1. 检查输入文件是否存在
    if not os.path.exists(INPUT_TEXT_FILE):
        print(f"错误：找不到输入的法律文书 {INPUT_TEXT_FILE}")
        print("请在脚本同目录下创建该文件，并填入法律文书原文。")
        return

    # 2. 读取法律文书原文
    try:
        with open(INPUT_TEXT_FILE, 'r', encoding='utf-8') as f:
            raw_legal_text = f.read()
        if not raw_legal_text.strip():
            print(f"错误：{INPUT_TEXT_FILE} 是空的。")
            return
    except Exception as e:
        print(f"读取 {INPUT_TEXT_FILE} 时发生错误: {e}")
        return

    # 3. 调用LLM提取数据
    extracted_data = call_lm_studio(raw_legal_text)

    if extracted_data:
        # 4. 如果成功提取，则写入Excel
        write_to_excel(extracted_data, OUTPUT_EXCEL_FILE)
    else:
        print("未能从LLM获取有效数据，已停止写入Excel。")


if __name__ == "__main__":
    main()