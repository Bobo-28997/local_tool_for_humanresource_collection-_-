# =====================================
# Streamlit App: 提成表多sheet自动审核 (App 3)
# 转换为 Tkinter 桌面应用
#
# 【Calamine 整合版】
# - [新] 所有 pd.read_excel/pd.ExcelFile 调用均已切换到 'calamine' 引擎
# - 包含 Streamlit -> Tkinter 的转换 (UI, 线程等)
# - 包含 pd.ExcelFile 的 'with' 语句修复
# - 移植了 App 3 特有的逻辑
# =====================================

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import unicodedata, re
import threading
import os
import traceback  # 用于更详细的错误日志


# =====================================
# 🛠️ 全局工具函数 (来自 Streamlit 脚本)
# =====================================
# (这些函数是纯 Python/Pandas 逻辑, 无需修改)

def normalize_text(val):
    if pd.isna(val):
        return ""
    s = str(val)
    s = re.sub(r'[\n\r\t ]+', '', s)
    s = s.replace('\u3000', '')
    return ''.join(unicodedata.normalize('NFKC', ch) for ch in s).lower().strip()


def normalize_num(val):
    if pd.isna(val):
        return None
    s = str(val).replace(",", "").strip()  # <--- 1. 不再替换 "%"
    if s in ["", "-", "nan"]:
        return None
    try:
        # 2. 在这里检查和处理 "%"
        if "%" in s:
            s = s.replace("%", "")
            return float(s) / 100
        return float(s)
    except:
        return s


def find_col(df_like, keyword, exact=False):
    key = keyword.strip().lower()
    columns = df_like.columns if hasattr(df_like, "columns") else df_like.index
    for col in columns:
        cname = str(col).strip().lower()
        if (exact and cname == key) or (not exact and key in cname):
            return col
    return None


def normalize_contract_key(series: pd.Series) -> pd.Series:
    s = series.astype(str)
    s = s.str.replace(r"\.0$", "", regex=True)
    s = s.str.strip()
    s = s.str.upper()
    s = s.str.replace('－', '-', regex=False)
    return s


def compare_series_vec(s_main, s_ref, compare_type='text', tolerance=0, multiplier=1):
    """
    (新) 向量化比较函数，复刻所有业务逻辑。
    """
    # 0. 识别 Merge 失败
    merge_failed_mask = s_ref.isna()

    # 1. 预处理空值
    main_is_na = pd.isna(s_main) | (s_main.astype(str).str.strip().isin(["", "nan", "None"]))
    ref_is_na = pd.isna(s_ref) | (s_ref.astype(str).str.strip().isin(["", "nan", "None"]))
    both_are_na = main_is_na & ref_is_na

    errors = pd.Series(False, index=s_main.index)

    # 2. 日期比较
    if compare_type == 'date':
        d_main = pd.to_datetime(s_main, errors='coerce').dt.normalize()
        d_ref = pd.to_datetime(s_ref, errors='coerce').dt.normalize()

        valid_dates_mask = d_main.notna() & d_ref.notna()
        date_diff_mask = (d_main != d_ref)
        errors = valid_dates_mask & date_diff_mask

        one_is_date_one_is_not = (d_main.notna() & d_ref.isna() & ~ref_is_na) | \
                                 (d_main.isna() & ~main_is_na & d_ref.notna())
        errors |= one_is_date_one_is_not

    # 3. 数值比较
    elif compare_type == 'num' or compare_type == 'rate' or compare_type == 'term':
        s_main_norm = s_main.apply(normalize_num)
        s_ref_norm = s_ref.apply(normalize_num)

        # 特殊：期限（乘数）
        if compare_type == 'term':
            s_ref_norm = pd.to_numeric(s_ref_norm, errors='coerce') * multiplier

        is_num_main = s_main_norm.apply(lambda x: isinstance(x, (int, float)))
        is_num_ref = s_ref_norm.apply(lambda x: isinstance(x, (int, float)))
        both_are_num = is_num_main & is_num_ref

        if both_are_num.any():
            diff = (s_main_norm[both_are_num] - s_ref_norm[both_are_num]).abs()
            errors.loc[both_are_num] = (diff > (tolerance + 1e-6))

        one_is_num_one_is_not = (is_num_main & ~is_num_ref & ~ref_is_na) | \
                                (~is_num_main & ~main_is_na & is_num_ref)
        errors |= one_is_num_one_is_not

    # 4. 文本比较
    else:  # compare_type == 'text'
        s_main_norm_text = s_main.apply(normalize_text)
        s_ref_norm_text = s_ref.apply(normalize_text)
        errors = (s_main_norm_text != s_ref_norm_text)

    # 5. 最终错误逻辑
    final_errors = errors & ~both_are_na
    lookup_failure_mask = merge_failed_mask & ~main_is_na
    final_errors = final_errors & ~lookup_failure_mask

    return final_errors


# =====================================
# 🖥️ Tkinter 应用主类
# =====================================

class AuditApp:
    def __init__(self, root):
        self.root = root
        # --- (【Calamine 修改】) ---
        self.root.title("📊 人事薪资表自动审核系统-3 二手车")
        self.root.geometry("800x700")

        self.style = ttk.Style()
        self.style.theme_use('clam')

        # --- 实例变量 ---
        self.uploaded_files = {}  # 存储文件路径
        self.output_dir = ""
        self.required_files = ["提成", "放款明细", "二次明细", "原表"]

        # --- GUI 布局 (与 App 2 相同) ---
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- 1. 文件选择区 ---
        input_frame = ttk.LabelFrame(main_frame, text="1. 输入文件", padding="10")
        input_frame.pack(fill=tk.X, expand=False)

        self.select_files_button = ttk.Button(input_frame, text=f"选择 {len(self.required_files)} 个 Excel 文件",
                                              command=self.select_files)
        self.select_files_button.pack(side=tk.LEFT, padx=(0, 10))

        self.file_status_label = ttk.Label(input_frame, text="尚未选择文件")
        self.file_status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- 2. 输出文件夹选择区 ---
        output_frame = ttk.LabelFrame(main_frame, text="2. 输出文件夹", padding="10")
        output_frame.pack(fill=tk.X, expand=False, pady=5)

        self.select_output_dir_button = ttk.Button(output_frame, text="选择报告保存位置",
                                                   command=self.select_output_dir)
        self.select_output_dir_button.pack(side=tk.LEFT, padx=(0, 10))

        self.output_dir_label = ttk.Label(output_frame, text="尚未选择文件夹")
        self.output_dir_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- 3. 执行区 ---
        run_frame = ttk.Frame(main_frame, padding="10")
        run_frame.pack(fill=tk.X, expand=False)

        self.run_button = ttk.Button(run_frame, text="🚀 开始审核", command=self.run_audit_thread, state=tk.DISABLED)
        self.run_button.pack(fill=tk.X, expand=True)

        # --- 4. 状态和进度区 ---
        status_frame = ttk.LabelFrame(main_frame, text="3. 审核状态", padding="10")
        status_frame.pack(fill=tk.X, expand=False, pady=5)

        self.status_label = ttk.Label(status_frame, text="等待开始...")
        self.status_label.pack(fill=tk.X, expand=True)

        self.progress_bar = ttk.Progressbar(status_frame, orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress_bar.pack(fill=tk.X, expand=True, pady=5)

        # --- 5. 日志区 ---
        log_frame = ttk.LabelFrame(main_frame, text="4. 运行日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.log_widget = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=20, state=tk.DISABLED)
        self.log_widget.pack(fill=tk.BOTH, expand=True)

        self.log_widget.tag_config('INFO', foreground='black')
        self.log_widget.tag_config('SUCCESS', foreground='green')
        self.log_widget.tag_config('WARNING', foreground='orange')
        self.log_widget.tag_config('ERROR', foreground='red', font=('Helvetica', '9', 'bold'))

    # =====================================
    # 🖥️ GUI 交互方法 (与 App 2 相同)
    # =====================================

    def _log(self, message, level='INFO'):
        """线程安全的日志更新"""
        self.root.after(0, self.update_log_widget, message, level)

    def update_log_widget(self, message, level):
        self.log_widget.config(state=tk.NORMAL)
        self.log_widget.insert(tk.END, f"{message}\n", level.upper())
        self.log_widget.see(tk.END)  # 自动滚动到底部
        self.log_widget.config(state=tk.DISABLED)

    def _update_status(self, text):
        """线程安全的状态标签更新"""
        self.root.after(0, self.status_label.config, {'text': text})

    def _update_progress(self, value):
        """线程安全的进度条更新 (value: 0.0 to 1.0)"""
        self.root.after(0, self.progress_bar.config, {'value': value * 100})

    def _set_gui_state(self, is_running):
        """线程安全地切换按钮状态"""
        self.root.after(0, self.toggle_buttons, is_running)

    def toggle_buttons(self, is_running):
        state = tk.DISABLED if is_running else tk.NORMAL
        self.select_files_button.config(state=state)
        self.select_output_dir_button.config(state=state)
        if not is_running and self.check_ready(silent=True):
            self.run_button.config(state=tk.NORMAL)
        else:
            self.run_button.config(state=tk.DISABLED)

    def check_ready(self, silent=False):
        """检查所有条件是否满足，以启用“开始审核”按钮"""
        ready = len(self.uploaded_files) == len(self.required_files) and self.output_dir
        if ready:
            self.run_button.config(state=tk.NORMAL)
            if not silent:
                self._log("✅ 所有文件和输出目录已就绪，可以开始审核。", "SUCCESS")
        else:
            self.run_button.config(state=tk.DISABLED)
        return ready

    def select_files(self):
        self.uploaded_files = {}  # 重置

        file_paths = filedialog.askopenfilenames(
            title=f"请选择所有 {len(self.required_files)} 个文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if not file_paths:
            self.file_status_label.config(text="未选择文件")
            self.check_ready()
            return

        found_count = 0
        missing_files = list(self.required_files)  # 复制一份

        for keyword in self.required_files:
            found_for_keyword = False
            for path in file_paths:
                filename = os.path.basename(path)
                if keyword in filename:
                    self.uploaded_files[keyword] = path
                    found_count += 1
                    if keyword in missing_files:
                        missing_files.remove(keyword)
                    found_for_keyword = True
                    break

        self.file_status_label.config(
            text=f"已选择 {found_count} / {len(self.required_files)} 个文件。"
        )

        if missing_files:
            self._log(f"⚠️ 仍缺少文件: {', '.join(missing_files)}", "WARNING")

        self.check_ready()

    def select_output_dir(self):
        dir_path = filedialog.askdirectory(title="选择报告保存的文件夹")
        if dir_path:
            self.output_dir = dir_path
            self.output_dir_label.config(text=f"将保存到: {dir_path}")
        else:
            self.output_dir = ""
            self.output_dir_label.config(text="尚未选择文件夹")
        self.check_ready()

    def run_audit_thread(self):
        """“开始审核”按钮的入口，启动一个新线程"""
        self._set_gui_state(is_running=True)
        self.log_widget.config(state=tk.NORMAL)
        self.log_widget.delete('1.0', tk.END)  # 清空日志
        self.log_widget.config(state=tk.DISABLED)

        self._log("🚀 审核任务开始...", "INFO")

        threading.Thread(target=self.run_audit_logic, daemon=True).start()

    # =====================================
    # 🧮 核心审核逻辑 (移植自 App 3)
    # =====================================

    def prepare_ref_df(self, df_list, required_cols_dict, prefix):
        """
        (新 V2) 预处理参考DF列表：合并、标准化Key、提取列、重命名
        """
        if not df_list or all(df is None for df in df_list):
            self._log(f"⚠️ {prefix} 数据列表为空，跳过预处理。", "WARNING")
            return pd.DataFrame(columns=['__KEY__'])

        try:
            df_concat = pd.concat([df for df in df_list if df is not None], ignore_index=True)
        except Exception as e:
            self._log(f"❌ 预处理 {prefix} 时合并失败: {e}", "ERROR")
            return pd.DataFrame(columns=['__KEY__'])

        contract_col_kw, contract_exact = required_cols_dict.get('合同', ('合同', False))
        contract_col = find_col(df_concat, contract_col_kw, exact=contract_exact)

        if not contract_col:
            self._log(
                f"⚠️ 在 {prefix} 参考表中未找到'合同'列 (关键字: '{contract_col_kw}', 精确: {contract_exact})，跳过此数据源。",
                "WARNING")
            return pd.DataFrame(columns=['__KEY__'])

        cols_to_extract = [contract_col]
        col_mapping = {}

        for std_name, (col_kw, is_exact) in required_cols_dict.items():
            if std_name == '合同': continue

            actual_col = find_col(df_concat, col_kw, exact=is_exact)

            if actual_col:
                cols_to_extract.append(actual_col)
                col_mapping[actual_col] = f"ref_{prefix}_{std_name}"
            else:
                self._log(f"⚠️ 在 {prefix} 参考表中未找到列 (关键字: '{col_kw}', 精确: {is_exact})", "WARNING")

        if len(cols_to_extract) == 1:
            self._log(f"⚠️ 在 {prefix} 参考表中未找到任何所需字段，跳过。", "WARNING")
            return pd.DataFrame(columns=['__KEY__'])

        std_df = df_concat[list(set(cols_to_extract))].copy()
        std_df['__KEY__'] = normalize_contract_key(std_df[contract_col])
        std_df = std_df.rename(columns=col_mapping)

        final_cols = ['__KEY__'] + list(col_mapping.values())

        final_cols_in_df = [col for col in final_cols if col in std_df.columns]
        std_df = std_df[final_cols_in_df]

        std_df = std_df.drop_duplicates(subset=['__KEY__'], keep='first')
        return std_df

    def audit_one_sheet_vec(self, tc_df, sheet_label, all_std_dfs, MAPPING):
        """
        (App 3 核心函数) 向量化审核单个 Sheet
        """
        contract_col_main = find_col(tc_df, "合同")
        if not contract_col_main:
            self._log(f"⚠️ {sheet_label}：未找到‘合同’列，跳过。", "WARNING")
            return 0, 0

        tc_df['__ROW_IDX__'] = tc_df.index
        tc_df['__KEY__'] = normalize_contract_key(tc_df[contract_col_main])

        merged_df = tc_df.copy()
        for std_df in all_std_dfs.values():
            if not std_df.empty:
                merged_df = pd.merge(merged_df, std_df, on='__KEY__', how='left')

        total_errors = 0
        errors_locations = set()  # 存储 (row_idx, col_name)
        row_has_error = pd.Series(False, index=merged_df.index)

        self._update_progress(0)

        for i, (main_kw, (src, ref_kw, tol, mult)) in enumerate(MAPPING.items()):

            exact_main = "期限" in main_kw or main_kw == "人员类型"
            main_col = find_col(merged_df, main_kw, exact=exact_main)
            if not main_col:
                continue

            self._update_status(f"{sheet_label} 审核进度：{i + 1}/{len(MAPPING)} - {main_kw}")

            s_main = merged_df[main_col]

            errors_mask = None
            if main_kw == "收益率":
                person_type_col = find_col(merged_df, "人员类型", exact=True)
                if not person_type_col:
                    continue

                s_ref_fk = merged_df.get('ref_fk_xirr')
                s_ref_orig = merged_df.get('ref_orig_年化nim')

                if s_ref_fk is None:
                    s_ref_fk = pd.Series(pd.NA, index=merged_df.index)

                s_ref_final = s_ref_fk.copy()

                if s_ref_orig is not None:
                    person_type_normalized = merged_df[person_type_col].apply(normalize_text)
                    mask_light_truck = (person_type_normalized == "轻卡")
                    s_ref_final.loc[mask_light_truck] = s_ref_orig.loc[mask_light_truck]

                errors_mask = compare_series_vec(s_main, s_ref_final, compare_type='rate', tolerance=tol)

            elif "日期" in main_kw or main_kw == "二次交接":
                ref_col_name = f"ref_{'ec' if src == '二次明细' else 'fk'}_{ref_kw}"
                s_ref = merged_df.get(ref_col_name)
                errors_mask = compare_series_vec(s_main, s_ref, compare_type='date')

            elif "期限" in main_kw:
                ref_col_name = f"ref_fk_{ref_kw}"
                s_ref = merged_df.get(ref_col_name)
                errors_mask = compare_series_vec(s_main, s_ref, compare_type='term', tolerance=tol, multiplier=mult)

            elif main_kw in ["租赁本金", "家访", "计算提成金额"]:
                ref_col_name = f"ref_fk_{ref_kw}"
                s_ref = merged_df.get(ref_col_name)
                errors_mask = compare_series_vec(s_main, s_ref, compare_type='num', tolerance=tol)

            else:  # 文本
                ref_col_name = f"ref_fk_{ref_kw}"
                s_ref = merged_df.get(ref_col_name)
                errors_mask = compare_series_vec(s_main, s_ref, compare_type='text')

            if errors_mask is not None and errors_mask.any():
                total_errors += errors_mask.sum()
                row_has_error |= errors_mask

                bad_indices = merged_df[errors_mask]['__ROW_IDX__']
                for idx in bad_indices:
                    errors_locations.add((idx, main_col))

            self._update_progress((i + 1) / len(MAPPING))

        self._update_status(f"{sheet_label} 比对完成，正在生成标注文件...")

        wb = Workbook()
        ws = wb.active
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        original_cols_list = list(tc_df.drop(columns=['__ROW_IDX__', '__KEY__']).columns)
        col_name_to_idx = {name: i + 1 for i, name in enumerate(original_cols_list)}

        for r in dataframe_to_rows(merged_df[original_cols_list], index=False, header=True):
            ws.append(r)

        for (row_idx, col_name) in errors_locations:
            if col_name in col_name_to_idx:
                excel_row = row_idx + 2
                excel_col = col_name_to_idx[col_name]
                ws.cell(excel_row, excel_col).fill = red_fill

        if contract_col_main in col_name_to_idx:
            contract_col_excel_idx = col_name_to_idx[contract_col_main]
            error_row_indices = merged_df[row_has_error]['__ROW_IDX__']
            for row_idx in error_row_indices:
                excel_row = row_idx + 2
                ws.cell(excel_row, contract_col_excel_idx).fill = yellow_fill

        save_path_full = os.path.join(self.output_dir, f"提成_{sheet_label}_审核标注版.xlsx")
        wb.save(save_path_full)
        self._log(f"📥 报告已保存: {save_path_full}", "SUCCESS")

        error_row_count = row_has_error.sum()

        if error_row_count > 0:
            try:
                df_errors_only = merged_df.loc[row_has_error, original_cols_list].copy()

                original_indices_with_error = merged_df.loc[row_has_error, '__ROW_IDX__']
                original_idx_to_new_excel_row = {
                    original_idx: new_row_num
                    for new_row_num, original_idx in enumerate(original_indices_with_error, start=2)
                }

                wb_err = Workbook()
                ws_err = wb_err.active

                for r in dataframe_to_rows(df_errors_only, index=False, header=True):
                    ws_err.append(r)

                for (original_row_idx, col_name) in errors_locations:
                    if original_row_idx in original_idx_to_new_excel_row:
                        new_row = original_idx_to_new_excel_row[original_row_idx]
                        if col_name in col_name_to_idx:
                            new_col = col_name_to_idx[col_name]
                            ws_err.cell(row=new_row, column=new_col).fill = red_fill

                contract_col_excel_idx = col_name_to_idx[contract_col_main]
                for new_row_num in original_idx_to_new_excel_row.values():
                    ws_err.cell(row=new_row_num, column=contract_col_excel_idx).fill = yellow_fill

                save_path_err = os.path.join(self.output_dir, f"提成_{sheet_label}_错误精简版.xlsx")
                wb_err.save(save_path_err)
                self._log(f"📥 错误精简版已保存: {save_path_err}", "SUCCESS")

            except Exception as e:
                self._log(f"❌ 生成“错误精简版”文件时出错: {e}", "ERROR")

        return total_errors, error_row_count

    def run_audit_logic(self):
        """
        这是在后台线程中运行的主函数。
        (App 3 逻辑)
        """
        try:
            # =====================================
            # 🛠️ 1. 文件路径准备
            # =====================================

            tc_file = self.uploaded_files.get("提成")
            fk_file = self.uploaded_files.get("放款明细")
            ec_file = self.uploaded_files.get("二次明细")
            original_file = self.uploaded_files.get("原表")

            if not all([tc_file, fk_file, ec_file, original_file]):
                self._log("❌ 内部错误：文件字典不完整。", "ERROR")
                return

            self._log("ℹ️ -------------------------------", "INFO")
            self.log_widget.insert(tk.END, f"ℹ️ 阶段 1/4: 正在读取 '提成' 文件...\n", 'INFO')

            # =====================================
            # 🛠️ 2. 读取主表 (提成)
            # =====================================
            tc_sheets = {}
            # --- (【Calamine 修改】) ---
            with pd.ExcelFile(tc_file, engine='calamine') as tc_xls:
                sheet_total = next((s for s in tc_xls.sheet_names if "总" in s), None)
                sheets_qk = [s for s in tc_xls.sheet_names if "轻卡" in s]
                sheets_zk = [s for s in tc_xls.sheet_names if "重卡" in s]

                tc_sheets = {
                    "总": [pd.read_excel(tc_xls, sheet_name=sheet_total)] if sheet_total else [],
                    "轻卡": [pd.read_excel(tc_xls, sheet_name=s) for s in sheets_qk],
                    "重卡": [pd.read_excel(tc_xls, sheet_name=s) for s in sheets_zk],
                }

            self._log(
                f"✅ 提成表已读取：总({len(tc_sheets['总'])})、轻卡({len(tc_sheets['轻卡'])})、重卡({len(tc_sheets['重卡'])})",
                "SUCCESS")
            self._log("ℹ️ -------------------------------", "INFO")
            self._log("ℹ️ 阶段 2/4: 正在读取并预处理参考文件...", "INFO")

            # =====================================
            # 🛠️ 3. 读取并预处理参考表
            # =====================================

            # --- 放款明细 (fk) ---
            fk_dfs_raw = []
            # --- (【Calamine 修改】) ---
            with pd.ExcelFile(fk_file, engine='calamine') as fk_xls:
                fk_sheet_names = [s for s in fk_xls.sheet_names if "潮掣" in s]
                self._log(f"ℹ️ 正在从 '放款明细' 加载 {len(fk_sheet_names)} 个 '潮掣' sheet...", "INFO")
                fk_dfs_raw = [pd.read_excel(fk_xls, sheet_name=s) for s in fk_sheet_names]

            fk_cols_needed = {
                '合同': ('合同', False),
                '放款日期': ('放款日期', False),
                '提报人员': ('提报人员', False),
                '城市经理': ('城市经理', False),
                '租赁本金': ('租赁本金', False),
                'xirr': ('xirr', False),
                '租赁期限/年': ('租赁期限/年', False),
                '家访': ('家访', False),
                '类型': ('类型', True),
                '净融资额': ('净融资额', False)  # (来自用户代码)
            }
            fk_std = self.prepare_ref_df(fk_dfs_raw, fk_cols_needed, "fk")

            # --- 二次明细 (ec) ---
            ec_dfs_raw = []
            # --- (【Calamine 修改】) ---
            with pd.ExcelFile(ec_file, engine='calamine') as ec_xls:
                self._log(f"ℹ️ 正在从 '二次明细' 加载 {len(ec_xls.sheet_names)} 个 sheet...", "INFO")
                ec_dfs_raw = [pd.read_excel(ec_xls, sheet_name=s) for s in ec_xls.sheet_names]

            ec_cols_needed = {'合同': ('合同', False), '出本流程时间': ('出本流程时间', False)}
            ec_std = self.prepare_ref_df(ec_dfs_raw, ec_cols_needed, "ec")

            # --- 原表 (original) ---
            self._log("ℹ️ 正在从 '原表' 加载数据...", "INFO")
            # --- (【Calamine 修改】) ---
            original_dfs_raw = [pd.read_excel(original_file, engine='calamine')]
            original_cols_needed = {'合同': ('合同', False), '年化nim': ('年化nim', False)}
            orig_std = self.prepare_ref_df(original_dfs_raw, original_cols_needed, "orig")

            all_std_dfs = {"fk": fk_std, "ec": ec_std, "orig": orig_std}
            self._log("✅ 所有参考文件已预处理完成。", "SUCCESS")

            # --- MAPPING ---
            MAPPING = {
                "放款日期": ("放款明细", "放款日期", 0, 1),
                "提报人员": ("放款明细", "提报人员", 0, 1),
                "城市经理": ("放款明细", "城市经理", 0, 1),
                "租赁本金": ("放款明细", "租赁本金", 0, 1),
                "收益率": ("放款明细", "xirr", 0.005, 1),
                "期限": ("放款明细", "租赁期限/年", 0.5, 12),
                "家访": ("放款明细", "家访", 0, 1),
                "人员类型": ("放款明细", "类型", 0, 1),
                "二次交接": ("二次明细", "出本流程时间", 0, 1),
                "计算提成金额": ("放款明细", "净融资额", 0, 1)  # (来自用户代码)
            }

            # =====================================
            # 🛠️ 4. 审核所有 sheet
            # =====================================
            self._log("ℹ️ -------------------------------", "INFO")
            self._log("ℹ️ 阶段 3/4: 正在执行主流程审核...", "INFO")
            results = {}
            for label, df_list in tc_sheets.items():
                if not df_list:
                    continue
                for i, df in enumerate(df_list, start=1):
                    tag = f"{label}{i if len(df_list) > 1 else ''}"
                    self._log(f"--- 📘 正在审核：{tag} ---", "INFO")
                    errs, rows = self.audit_one_sheet_vec(df, tag, all_std_dfs, MAPPING)
                    results[tag] = (errs, rows)

            # =====================================
            # 🛠️ 5. 反向漏填检查
            # =====================================
            self._log("ℹ️ -------------------------------", "INFO")
            self._log("ℹ️ 阶段 4/4: 正在执行反向漏填检查...", "INFO")

            contracts_total = set()
            if tc_sheets["总"]:
                df_total = tc_sheets["总"][0]
                col = find_col(df_total, "合同", exact=False)
                if col is not None:
                    contracts_total = set(normalize_contract_key(df_total[col].dropna()))

            contracts_fk = set(fk_std['__KEY__'].dropna())
            missing_contracts = sorted(list(contracts_fk - contracts_total))

            if missing_contracts:
                self._log(f"⚠️ 发现 {len(missing_contracts)} 个合同号存在于放款明细中，但未出现在提成表‘总’sheet中",
                          "WARNING")

                wb_miss = Workbook()
                ws_miss = wb_miss.active
                ws_miss.cell(1, 1, "漏填合同号")
                for r, contract in enumerate(missing_contracts, start=2):
                    ws_miss.cell(r, 1, contract)

                save_path_missing = os.path.join(self.output_dir, "提成_漏填合同号_基于放款明细_潮掣.xlsx")
                wb_miss.save(save_path_missing)
                self._log(f"📥 漏填合同列表已保存: {save_path_missing}", "SUCCESS")
            else:
                self._log("✅ 未发现漏填合同号（基于放款明细-潮掣）。", "SUCCESS")

            # =====================================
            # 🛠️ 6. 结果汇总 (替换下载区)
            # =====================================
            self._log("ℹ️ -------------------------------", "INFO")
            self._log("📤 审核结果摘要 (文件已自动保存至输出文件夹):", "INFO")

            total_all_errors = 0
            for tag, (errs, rows) in results.items():
                self._log(f"📘 **{tag}**：发现 {errs} 个错误，共 {rows} 行异常")
                total_all_errors += errs

            if total_all_errors == 0:
                self._log("✅ 恭喜！所有文件中未发现任何错误。", "SUCCESS")

            self._log("🎉🎉🎉 所有审核任务已完成！", "SUCCESS")

        except Exception as e:
            self._log(f"❌❌❌ 发生未捕获的严重错误: {e}", "ERROR")
            self._log(traceback.format_exc(), "ERROR")
            self.root.after(0, messagebox.showerror, "严重错误", f"发生未捕获的错误: \n{e}")

        finally:
            self._set_gui_state(is_running=False)
            self._update_status("审核完成。")
            self._update_progress(0)
            self.root.after(0, messagebox.showinfo, "任务完成", "审核已全部完成！")


# =====================================
# 🚀 启动应用
# =====================================
if __name__ == "__main__":
    root = tk.Tk()
    app = AuditApp(root)
    root.mainloop()